import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
old_path =r"C:\TMP\FaultMatrix_HDT (6).xls"
new_path =r"C:\TMP\FaultMatrix_HDT.xls"
old_excel = pd.read_excel(old_path) # Read the old  Excel files into pandas dataframes
new_excel = pd.read_excel(new_path) # Read the new Excel files into pandas dataframes

# Define the column to use as the common identifier
co_input = "ID"

# Get the list of columns in the old and new Excel files
col = list(old_excel.columns)
col2 = list(new_excel.columns)

# Define lists to store the columns that are in one file but not the other
main_col, index_list, index_list2, not_list, not_list2, new_col, old_col  = [], [], [], [], [], [], []

# Check if the columns in the old and new Excel files are the same
if col == col2:
    print("yes")
    main_col = col
else:
    # If the columns are not the same, determine which columns are in one file but not the other
    if len(col)>len(col2):
        main_col = col
        for i in range(0, len(col2)):
            k = 0
            for j in range(0, len(main_col)):
                if(main_col[j] == col2[i]):
                    k = 1
                    index_list2.append(i)
            if k!=1:
                not_list2.append(i)
                new_col.append(col2[i])
        for i in range(0, len(main_col)):
            k=0
            for j in range(0, len(col2)):
                if(main_col[i] == col2[j]):
                    k=1
                    index_list.append(i)
            if k!=1:
                not_list.append(i)
                old_col.append(col[i])
    else:
        main_col = col2
        for i in range(0, len(col)):
            k=0
            for j in range(0, len(main_col)):
                if(main_col[j] == col[i]):
                    k=1
                    index_list.append(i)
            if k!=1:
                not_list.append(i)
                old_col.append(col[i])
        for i in range(0, len(main_col)):
            k=0
            for j in range(0, len(col)):
                if(main_col[i]==col[j]):
                    k=1
                    index_list2.append(i)
            if k!=1:
                not_list2.append(i)
                new_col.append(col2[i])

# Determine which dataframe has more rows and use that as the first dataframe
l1=0
l2=0
first_id=[]
second_id=[]
#if len(old_excel)>len(new_excel):
l1=len(old_excel)
l2=len(new_excel)
first_id=list(old_excel[co_input])
second_id=list(new_excel[co_input])

# Define lists to store the modified rows
emp=[]
id_list=[]

# Iterate over the rows in the first and second dataframe
for i in range(0, l1):
    temp=[]
    for j in range(0, l2):
        if first_id[i]==second_id[j]:
            print(first_id[i],second_id[j],i)
            il1=list(old_excel.iloc[i])
            il2=list(new_excel.iloc[j])
            if il1==il2:
                temp = list(old_excel.iloc[i])
            else:
                if col == col2:
                    for k in range(0, len(il1)):
                        if il1[k] == il2[k] or il1[k] is np.nan and il2[k] is np.nan:
                            temp.append(il1[k])
                        else:
                            temp.append("old--->"+str(il1[k])+"\n"+"new--->"+str(il2[k]))
                else:
                    check=[]
                    inc=0
                    m=0

                    # Iterate over the columns in the first dataframe
                    for k in index_list:
                        if inc<=len(index_list2):
                            m=index_list2[inc]
                            inc+=1
                        if il1[k]==il2[m]:
                            temp.append(il1[k])
                        else:
                            # If the values are not the same, add the old and new values to the list 
                            if il1[k] is np.nan and il2[m] is np.nan:
                                s=""
                                temp.append(s)
                            else:
                                s="old--->"+str(il1[k])+"\n"+"new--->"+str(il2[m])
                                temp.append(s)

                    # Add any columns that are in the first dataframe but not the second dataframe
                    for k in not_list:
                        temp.insert(k, il1[k])

                    # Add any columns that are in the second dataframe but not the first dataframe
                    for m in range(0, len(not_list2)):
                        temp.append(il2[not_list2[m]])
            id_list.append(first_id[i])

    if(len(temp)!=0):
        emp.append(temp)
for i in new_col:
    if len(col)<len(col2):
        main_col[len(main_col)-1] ='new_ex--->'+i
    else:
        main_col.append("new_ex--->"+i)
    #main_col.append(il2[not_list2]) + 'new_ex--->'+new_col[i]
    #main_col.append('new_ex--->'+new_col[i])
new_ex = pd.DataFrame(emp,columns=main_col)
new_ex[co_input]=id_list
new_id=list(new_ex[co_input])

# Add any rows that are in the second dataframe but not the first dataframe
for i in range(0, len(first_id)):
    temp=[]
    new_temp=[]
    if first_id[i] not in new_id:
        temp.append(list(old_excel.iloc[i]))
        for t in range(0, len(temp[0])):
            new_temp.append('old--->'+str(temp[0][t]))
        for m in range(0, len(not_list2)):
            new_temp.append("")
        emp.append(new_temp)

# Add any rows that are in the first dataframe but not the second dataframe
for i in range(0, len(second_id)):
    temp=[]
    new_temp=[]
    if second_id[i] not in new_id:
        temp.append(list(new_excel.iloc[i]))
        for t in range(0, len(temp[0])):
            new_temp.append('new--->'+str(temp[0][t]))
        for k in not_list:
            new_temp.insert(k,"")
        emp.append(new_temp)
new_ex=pd.DataFrame(emp,columns=main_col)
for i in range(0, len(old_col)):
    new_ex.rename(columns={old_col[i]:'old_ex--->'+old_col[i]},inplace=True)
sp = old_path.split("\\")
newpath=""
for i in range(0, len(sp)-1):
    newpath = newpath+sp[i]+"\\"
newpath = newpath+"combined.xlsx"
new_ex.to_excel(newpath, index=None)
# Save the new dataframe to an Excel file
#new_ex.to_excel(r"C:\data\FaultMatrix\Data\SF4_1_0-101\combined_comparision.xlsx", header=True, index=None)

# Load the Excel file
wb = openpyxl.load_workbook(newpath)

# Select the desired sheet
ws = wb['Sheet1']

# Define the red fill color
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Iterate over the cells in the sheet
for row in ws.iter_rows():
    for cell in row:
        # Check if the cell value contains "-->"
        if cell.value and "-->" in str(cell.value):
            # Apply the red fill color to the cell
            cell.fill = red_fill

# Save the modified Excel file
wb.save(newpath)






